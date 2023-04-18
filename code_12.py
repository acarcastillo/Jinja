from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
from jinja2 import Environment, FileSystemLoader
import csv

# Load the Jinja2 template from a file
env = Environment(loader=FileSystemLoader('.'))
template = env.get_template('template.html')

# Render the template with some data
data = {'name': 'John Doe', 'age': 42}
html = template.render(data)

# Parse the HTML with BeautifulSoup
soup = BeautifulSoup(html, 'html.parser')

# Find the table you want to export (in this example, the first table)
table = soup.find('table')

# Extract the table headers into a list
headers = [th.text for th in table.find_all('th')]

# Extract the table rows into a list of lists
rows = []
for row in table.find_all('tr'):
    rows.append([cell.text for cell in row.find_all('td')])

# Create a new workbook
wb = Workbook()
ws = wb.active

# Add the headers as the first row of the worksheet
ws.append(headers)

# Add the data to the worksheet
for row in rows:
    ws.append(row)

# Add some formatting to the worksheet
for col in ws.columns:
    letter = get_column_letter(col[0].column)
    ws.column_dimensions[letter].width = 20

# Save the workbook as an Excel file
wb.save('output.xlsx')
